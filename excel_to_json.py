import openpyxl
import json
import os
import sys

# Load the Excel file (from first argument or default)
excel_path = sys.argv[1] if len(sys.argv) > 1 else "Bay_Area_Properties_50.xlsx"
output_path = sys.argv[2] if len(sys.argv) > 2 else "properties.json"

print(f"📂 Reading: {excel_path}")
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

properties = []

# Iterate through rows starting from row 3 (headers in row 2)
# Stop when we encounter a row with no ID (empty row)
for row_idx in range(3, ws.max_row + 1):
    # Break if we hit an empty row
    prop_id = ws.cell(row=row_idx, column=1).value
    if prop_id is None:
        break
    address = ws.cell(row=row_idx, column=3).value
    city = ws.cell(row=row_idx, column=4).value
    price = ws.cell(row=row_idx, column=5).value
    status = ws.cell(row=row_idx, column=6).value
    beds = ws.cell(row=row_idx, column=7).value
    baths = ws.cell(row=row_idx, column=8).value
    sqft = ws.cell(row=row_idx, column=9).value
    prop_type = ws.cell(row=row_idx, column=11).value
    year_built = ws.cell(row=row_idx, column=12).value
    hoa = ws.cell(row=row_idx, column=13).value
    agents_raw = ws.cell(row=row_idx, column=14).value
    licenses_raw = ws.cell(row=row_idx, column=15).value
    mls_num = ws.cell(row=row_idx, column=16).value
    description = ws.cell(row=row_idx, column=18).value
    parking = ws.cell(row=row_idx, column=30).value
    lot_size = ws.cell(row=row_idx, column=33).value
    
    # Clean up address for folder matching (address is already the folder name)
    folder_name = address if address else f"property-{int(prop_id)}"
    
    # Generate image paths (assuming 5 images per property: 01.jpg through 05.jpg)
    images = [
        f"/images/{folder_name}/01.jpg",
        f"/images/{folder_name}/02.jpg",
        f"/images/{folder_name}/03.jpg",
        f"/images/{folder_name}/04.jpg",
        f"/images/{folder_name}/05.jpg"
    ]
    
    # Split agents by "|" delimiter
    agents = []
    if agents_raw and agents_raw != "-":
        agents = [a.strip() for a in str(agents_raw).split("|")]
    
    # Split licenses by "|" delimiter
    agent_licenses = []
    if licenses_raw and licenses_raw != "-":
        agent_licenses = [l.strip() for l in str(licenses_raw).split("|")]
    
    # Convert numeric values
    try:
        price = int(float(price)) if price else None
        beds = int(float(beds)) if beds else None
        baths = int(float(baths)) if baths else None
        sqft = int(float(sqft)) if sqft else None
        year_built = int(float(year_built)) if year_built else None
        hoa = int(float(hoa)) if hoa and hoa != "-" else None
        mls_num = int(float(mls_num)) if mls_num else None
    except:
        pass
    
    # Build property object
    property_obj = {
        "id": int(prop_id) if prop_id else None,
        "address": address,
        "city": city,
        "price": price,
        "status": status,
        "beds": beds,
        "baths": baths,
        "sqft": sqft,
        "type": prop_type,
        "yearBuilt": year_built,
        "hoaMonth": hoa,
        "agents": agents,
        "agentLicenses": agent_licenses,
        "mlsNumber": mls_num,
        "description": description,
        "parking": parking,
        "lotSize": lot_size,
        "images": images
    }
    
    properties.append(property_obj)

# Write to JSON file
with open(output_path, "w") as f:
    json.dump(properties, f, indent=2)

print(f"✅ Conversion complete!")
print(f"📄 Output: {output_path}")
print(f"🏠 Total properties: {len(properties)}")
print(f"\n📋 Sample (first property):")
print(json.dumps(properties[0], indent=2))
